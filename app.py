from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from pypinyin import Style, lazy_pinyin


BASE_DIR = Path(__file__).resolve().parent
WORK_DIR = BASE_DIR / "work"
DIST_DIR = BASE_DIR / "dist"
OUTPUT_DIR = BASE_DIR / "output"
DAT_FILENAMES = {
    "custom_phrase": "UserDefinedPhrase.dat",
    "self_study": "SelfStudyPhrase.dat",
}
GENERATION_MODES = {
    "custom_phrase": {
        "label": "用户自定义短语",
        "output_format": "win10mspy",
        "code_name": "拼音首字母",
    },
    "self_study": {
        "label": "自学习词汇",
        "output_format": "win10mspyss",
        "code_name": "全拼编码",
    },
}
FALLBACK_CONVERTER_PATH = BASE_DIR / "imewlconverter_linux" / "publish" / "linux-x64" / "ImeWlConverterCmd"
DEFAULT_CONVERTER_PATH = os.environ.get("IMEWLCONVERTER_PATH", "").strip() or (
    str(FALLBACK_CONVERTER_PATH) if FALLBACK_CONVERTER_PATH.exists() else ""
)
NAME_HEADER_CANDIDATES = {"姓名", "name", "names"}


app = Flask(__name__)


class AppError(Exception):
    pass


@dataclass(frozen=True)
class NameRecord:
    name: str
    code: str


def ensure_runtime_dirs() -> None:
    for path in (WORK_DIR, DIST_DIR, OUTPUT_DIR):
        path.mkdir(parents=True, exist_ok=True)


def is_cjk(char: str) -> bool:
    code = ord(char)
    return (
        0x3400 <= code <= 0x4DBF
        or 0x4E00 <= code <= 0x9FFF
        or 0xF900 <= code <= 0xFAFF
    )


def normalize_name(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[^\u3400-\u9fffA-Za-z0-9]", "", text)
    return text


def build_initials(name: str) -> str:
    pieces = lazy_pinyin(name, style=Style.FIRST_LETTER, errors="keep")
    result: list[str] = []
    for piece in pieces:
        if not piece:
            continue
        for char in piece:
            if char.isalnum():
                result.append(char.lower())
    return "".join(result)


def build_full_pinyin(name: str) -> str:
    pieces = lazy_pinyin(name, style=Style.NORMAL, errors="ignore")
    return " ".join(piece.lower() for piece in pieces if piece)


def normalize_mode(mode: str) -> str:
    mode = (mode or "custom_phrase").strip()
    if mode not in GENERATION_MODES:
        raise AppError("生成模式不支持")
    return mode


def parse_manual_names(text: str) -> list[str]:
    return [line for line in (text or "").splitlines()]


def decode_text_bytes(payload: bytes) -> str:
    encodings = ("utf-8-sig", "utf-8", "gb18030", "gbk")
    for encoding in encodings:
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise AppError("上传文件无法识别编码，请保存为 UTF-8、GBK 或 GB18030 后重试")


def read_txt_names(file_bytes: bytes) -> list[str]:
    return decode_text_bytes(file_bytes).splitlines()


def read_csv_names(file_bytes: bytes) -> list[str]:
    text = decode_text_bytes(file_bytes)
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []

    first_row = [cell.strip().lower() for cell in rows[0]]
    target_index = 0
    has_header = False
    for idx, value in enumerate(first_row):
        if value in NAME_HEADER_CANDIDATES:
            target_index = idx
            has_header = True
            break

    start = 1 if has_header else 0
    names: list[str] = []
    for row in rows[start:]:
        if target_index < len(row):
            names.append(row[target_index])
    return names


def read_xlsx_names(file_bytes: bytes) -> list[str]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        values = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not values:
        return []

    first_row = [str(cell).strip().lower() if cell is not None else "" for cell in values[0]]
    target_index = 0
    start = 0
    for idx, value in enumerate(first_row):
        if value in NAME_HEADER_CANDIDATES:
            target_index = idx
            start = 1
            break

    names: list[str] = []
    for row in values[start:]:
        if target_index < len(row):
            cell = row[target_index]
            if cell is not None:
                names.append(str(cell))
    return names


def extract_uploaded_names(filename: str, file_bytes: bytes) -> list[str]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".txt":
        return read_txt_names(file_bytes)
    if suffix == ".csv":
        return read_csv_names(file_bytes)
    if suffix == ".xlsx":
        return read_xlsx_names(file_bytes)
    raise AppError("仅支持上传 txt、csv、xlsx 文件")


def build_records(raw_names: Iterable[str], mode: str) -> list[NameRecord]:
    records: list[NameRecord] = []
    seen: set[tuple[str, str]] = set()
    for raw_name in raw_names:
        name = normalize_name(str(raw_name))
        if not name:
            continue
        code = build_initials(name) if mode == "custom_phrase" else build_full_pinyin(name)
        if not code:
            continue
        key = (name, code)
        if key in seen:
            continue
        seen.add(key)
        records.append(NameRecord(name=name, code=code))
    return records


def validate_records(payload: list[dict], mode: str) -> list[NameRecord]:
    if not payload:
        raise AppError("未读取到有效姓名")

    records: list[NameRecord] = []
    seen: set[tuple[str, str]] = set()
    for item in payload:
        name = normalize_name(str(item.get("name", "")))
        raw_code = str(item.get("code", "")).strip().lower()
        if mode == "custom_phrase":
            code = re.sub(r"[^a-z0-9]", "", raw_code)
        else:
            code = re.sub(r"[^a-z'\s]", "", raw_code)
            code = re.sub(r"\s+", " ", code.replace("'", " ")).strip()
        if not name:
            continue
        if not code:
            raise AppError(f"“{name}” 的编码为空，请先修正后再生成")
        if mode == "self_study" and " " not in code and len(name) > 1:
            raise AppError(f"“{name}” 的全拼编码请用空格分隔，例如：zhang san")
        key = (name, code)
        if key in seen:
            continue
        seen.add(key)
        records.append(NameRecord(name=name, code=code))

    if not records:
        raise AppError("未读取到有效姓名")
    return records


def normalize_converter_path(raw_path: str) -> str:
    path = (raw_path or "").strip().strip('"').strip("'")
    windows_drive_match = re.match(r"^([A-Za-z]):[\\/](.*)$", path)
    if windows_drive_match:
        drive = windows_drive_match.group(1).lower()
        rest = windows_drive_match.group(2).replace("\\", "/")
        return f"/mnt/{drive}/{rest}"
    return path


def resolve_converter_command(converter_path: str) -> list[str]:
    raw = normalize_converter_path(converter_path or DEFAULT_CONVERTER_PATH)
    if not raw:
        raise AppError("ImeWlConverter 路径不存在")

    candidate = Path(raw)
    resolved = candidate if candidate.is_absolute() else (BASE_DIR / candidate).resolve()
    if resolved.exists():
        if resolved.suffix.lower() == ".dll":
            dotnet = shutil.which("dotnet")
            if not dotnet:
                raise AppError("检测到 DLL 版本转换器，但当前环境找不到 dotnet 命令")
            return [dotnet, str(resolved)]
        return [str(resolved)]

    executable = shutil.which(raw)
    if executable:
        return [executable]

    raise AppError("ImeWlConverter 路径不存在")


def build_converter_invocation(
    command: list[str],
    input_path: Path,
    output_path: Path,
    mode: str,
) -> list[str]:
    output_format = GENERATION_MODES[mode]["output_format"]
    executable = command[0]
    executable_name = Path(executable).name.lower()
    if executable_name == "imewlconvertercmd":
        return [*command, f"-i:sgpy", str(input_path), f"-o:{output_format}", str(output_path)]
    return [*command, "-i", "sgpy", "-o", output_format, "-O", str(output_path), str(input_path)]


def prepare_executable_command(command: list[str]) -> list[str]:
    executable = Path(command[0])
    if executable.name == "ImeWlConverterCmd" and executable.exists():
        temp_copy = Path("/tmp") / executable.name
        shutil.copy2(executable, temp_copy)
        temp_copy.chmod(0o755)
        return [str(temp_copy), *command[1:]]
    return command


def extract_real_dat_path(dat_path: Path) -> Path | None:
    if not dat_path.exists():
        return None

    raw = dat_path.read_bytes()
    text = None
    for encoding in ("utf-16", "utf-8", "gb18030", "gbk"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if not text:
        return None

    match = re.search(r"词库文件在：(.+)", text)
    if not match:
        return None

    candidate = Path(match.group(1).strip())
    return candidate if candidate.exists() else None


def split_record_code(record: NameRecord, mode: str) -> list[str]:
    if mode == "custom_phrase":
        return list(re.sub(r"[^a-z0-9]", "", record.code.lower()))
    return [part for part in re.split(r"['\s]+", record.code.lower()) if part]


def parse_sgpy_lines(text: str, mode: str) -> list[NameRecord]:
    records: list[NameRecord] = []
    seen: set[tuple[str, str]] = set()
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("'") or " " not in line:
            continue
        raw_code, raw_name = line.split(" ", 1)
        name = normalize_name(raw_name)
        code_parts = [part for part in raw_code.split("'") if part]
        if not name or not code_parts:
            continue
        code = "".join(part[0] for part in code_parts) if mode == "custom_phrase" else " ".join(code_parts)
        key = (name, code)
        if key in seen:
            continue
        seen.add(key)
        records.append(NameRecord(name=name, code=code))
    return records


def detect_dat_mode(file_bytes: bytes) -> str | None:
    if file_bytes.startswith(b"mschxudp"):
        return "custom_phrase"
    if file_bytes.startswith(bytes.fromhex("55 aa 88 81")):
        return "self_study"
    return None


def write_intermediate_file(records: list[NameRecord], mode: str) -> Path:
    normalized_path = WORK_DIR / "names_normalized.txt"
    lines = ["'" + "'".join(split_record_code(record, mode)) + f" {record.name}" for record in records]
    normalized_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return normalized_path


def generate_dat(records: list[NameRecord], converter_path: str, mode: str) -> dict:
    ensure_runtime_dirs()

    command = prepare_executable_command(resolve_converter_command(converter_path))
    normalized_path = write_intermediate_file(records, mode)
    dat_path = DIST_DIR / DAT_FILENAMES[mode]
    stdout_path = WORK_DIR / "converter_stdout.log"
    stderr_path = WORK_DIR / "converter_stderr.log"
    request_path = OUTPUT_DIR / "last_request.json"

    request_path.write_text(
        json.dumps(
            {
                "mode": mode,
                "records": [record.__dict__ for record in records],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    if dat_path.exists():
        dat_path.unlink()

    full_command = build_converter_invocation(command, normalized_path, dat_path, mode)

    completed = subprocess.run(
        full_command,
        cwd=BASE_DIR,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    stdout_path.write_text(completed.stdout or "", encoding="utf-8")
    stderr_path.write_text(completed.stderr or "", encoding="utf-8")

    if completed.returncode != 0:
        raise AppError("调用转换器失败，请检查转换器是否可运行")
    redirected_dat_path = extract_real_dat_path(dat_path)
    if redirected_dat_path:
        shutil.copyfile(redirected_dat_path, dat_path)
    if not dat_path.exists():
        raise AppError("DAT 文件生成失败，未找到输出文件")
    if dat_path.stat().st_size == 0:
        raise AppError("DAT 文件生成失败，输出为空文件")

    return {
        "message": f"{GENERATION_MODES[mode]['label']} DAT 文件生成成功",
        "download_url": f"/download/dat/{mode}",
        "dat_size": dat_path.stat().st_size,
        "command": full_command,
        "filename": DAT_FILENAMES[mode],
    }


def inspect_dat_file(file_bytes: bytes, filename: str, converter_path: str, mode: str) -> dict:
    ensure_runtime_dirs()

    suffix = Path(filename or "").suffix.lower()
    if suffix != ".dat":
        raise AppError("仅支持查看 dat 文件")

    detected_mode = detect_dat_mode(file_bytes)
    if detected_mode:
        mode = detected_mode

    command = prepare_executable_command(resolve_converter_command(converter_path))
    input_path = WORK_DIR / "inspect_input.dat"
    output_path = WORK_DIR / "inspect_output_sgpy.txt"
    stdout_path = WORK_DIR / "inspect_stdout.log"
    stderr_path = WORK_DIR / "inspect_stderr.log"

    input_path.write_bytes(file_bytes)
    if output_path.exists():
        output_path.unlink()

    input_format = GENERATION_MODES[mode]["output_format"]
    executable_name = Path(command[0]).name.lower()
    if executable_name == "imewlconvertercmd":
        full_command = [*command, f"-i:{input_format}", str(input_path), "-o:sgpy", str(output_path)]
    else:
        full_command = [*command, "-i", input_format, "-o", "sgpy", "-O", str(output_path), str(input_path)]

    completed = subprocess.run(
        full_command,
        cwd=BASE_DIR,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )

    stdout_path.write_text(completed.stdout or "", encoding="utf-8")
    stderr_path.write_text(completed.stderr or "", encoding="utf-8")

    if completed.returncode != 0:
        raise AppError("解析 DAT 失败，请确认文件类型和生成模式是否匹配")
    if not output_path.exists() or output_path.stat().st_size == 0:
        raise AppError("解析 DAT 失败，未生成可读内容")

    text = decode_text_bytes(output_path.read_bytes())
    records = parse_sgpy_lines(text, mode)
    if not records:
        raise AppError("DAT 文件中未解析到有效词条")

    return {
        "records": [record.__dict__ for record in records],
        "summary": {
            "count": len(records),
            "source_file": filename,
            "mode": mode,
            "code_name": GENERATION_MODES[mode]["code_name"],
            "detected_mode": detected_mode,
        },
    }


@app.get("/")
def index():
    return render_template("index.html", default_converter_path=DEFAULT_CONVERTER_PATH)


@app.post("/api/parse")
def parse_input():
    try:
        mode = normalize_mode(request.form.get("mode", "custom_phrase"))
        manual_names = parse_manual_names(request.form.get("manual_names", ""))
        upload = request.files.get("input_file")
        file_names: list[str] = []
        source_name = ""
        if upload and upload.filename:
            source_name = upload.filename
            file_names = extract_uploaded_names(upload.filename, upload.read())

        records = build_records([*manual_names, *file_names], mode)
        if not records:
            raise AppError("未读取到有效姓名")

        return jsonify(
            {
                "records": [record.__dict__ for record in records],
                "summary": {
                    "count": len(records),
                    "source_file": source_name,
                    "mode": mode,
                    "code_name": GENERATION_MODES[mode]["code_name"],
                },
            }
        )
    except AppError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        return jsonify({"error": "解析失败，请检查输入内容或文件格式"}), 500


@app.post("/api/generate")
def generate():
    try:
        payload = request.get_json(silent=True) or {}
        mode = normalize_mode(str(payload.get("mode", "custom_phrase")))
        records = validate_records(payload.get("records", []), mode)
        result = generate_dat(records, str(payload.get("converter_path", "")), mode)
        return jsonify(result)
    except AppError as exc:
        return jsonify({"error": str(exc)}), 400
    except FileNotFoundError:
        return jsonify({"error": "ImeWlConverter 路径不存在"}), 400
    except Exception:
        return jsonify({"error": "DAT 生成失败，请检查转换器日志"}), 500


@app.post("/api/inspect-dat")
def inspect_dat():
    try:
        mode = normalize_mode(request.form.get("mode", "custom_phrase"))
        upload = request.files.get("dat_file")
        if not upload or not upload.filename:
            raise AppError("请先选择 DAT 文件")
        result = inspect_dat_file(
            upload.read(),
            upload.filename,
            request.form.get("converter_path", ""),
            mode,
        )
        return jsonify(result)
    except AppError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        return jsonify({"error": "解析 DAT 失败，请检查转换器日志"}), 500


@app.get("/download/dat/<mode>")
def download_dat(mode: str):
    try:
        mode = normalize_mode(mode)
    except AppError as exc:
        return jsonify({"error": str(exc)}), 400
    dat_path = DIST_DIR / DAT_FILENAMES[mode]
    if not dat_path.exists() or dat_path.stat().st_size == 0:
        return jsonify({"error": "DAT 文件不存在，请先生成"}), 404
    return send_file(dat_path, as_attachment=True, download_name=DAT_FILENAMES[mode])


if __name__ == "__main__":
    ensure_runtime_dirs()
    app.run(host="0.0.0.0", port=5000, debug=True)
